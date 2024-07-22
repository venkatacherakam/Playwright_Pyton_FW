import pytest
import pandas as pd
import openpyxl


def get_data_from_excel(path, sheet_name="Sheet1"):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    for r in range(2, total_rows + 1):
        row_list = []
        for c in range(1, total_cols + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
            final_list.append(row_list)
    return final_list



def load_data_from_excel(path, sheet_name="Sheet1"):
    df = pd.read_excel(path, sheet_name)
    return df
